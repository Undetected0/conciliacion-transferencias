from io import BytesIO

import openpyxl
import pandas as pd

from config import HOJAS_OBJETIVO, HEADER_ROW_CONGLOMERADO, HEADER_ROW_REVISION
from utils import normalizar_columnas, encontrar_columna


def _leer_hoja(file_bytes, sheet_name, header_row):
    """Lee una hoja con openpyxl en modo read_only (streaming). Más rápido y
    eficiente en memoria que el modo normal, sin pérdida de precisión numérica."""
    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)

    # Avanzar hasta la fila de encabezado
    for _ in range(header_row):
        next(rows)
    headers = [str(c) if c is not None else "" for c in next(rows)]

    data = [[str(c) if c is not None else "" for c in row] for row in rows]

    wb.close()
    return pd.DataFrame(data, columns=headers)


def leer_archivo_conglomerado(file):
    dfs = []
    logs = []
    file_bytes = file.getvalue()

    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True)
    sheet_names = wb.sheetnames
    wb.close()

    logs.append(f"Procesando conglomerado: {file.name}")

    for hoja in HOJAS_OBJETIVO:
        if hoja in sheet_names:
            df = _leer_hoja(file_bytes, hoja, HEADER_ROW_CONGLOMERADO)
            df = df.dropna(how="all")
            df = normalizar_columnas(df)

            col_match = encontrar_columna(df, "Identificador de la Transaccion CCE Online")
            col_estado = encontrar_columna(df, "Estado")
            df = df[[col_match, col_estado]].copy()

            dfs.append(df)
            logs.append(f"{file.name} | {hoja}: {len(df)} filas")
        else:
            logs.append(f"{file.name} | {hoja}: hoja no encontrada")

    return dfs, logs


def consolidar_conglomerado(files):
    all_dfs = []
    logs = []

    for f in files:
        dfs, logs_arch = leer_archivo_conglomerado(f)
        all_dfs.extend(dfs)
        logs.extend(logs_arch)

    if not all_dfs:
        return None, logs

    df = pd.concat(all_dfs, ignore_index=True)
    df = df.dropna(how="all")
    return df, logs


def leer_revision(file):
    file_bytes = file.getvalue()

    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True)
    sheet_name = wb.sheetnames[0]
    wb.close()

    df = _leer_hoja(file_bytes, sheet_name, HEADER_ROW_REVISION)
    df = df.dropna(how="all")
    df = normalizar_columnas(df)
    return df


def consolidar_revision(files):
    dfs = []
    logs = []

    for f in files:
        df = leer_revision(f)
        df["archivo_revision_origen"] = f.name
        dfs.append(df)
        logs.append(f"{f.name}: {len(df)} filas")

    if not dfs:
        return None, logs

    df = pd.concat(dfs, ignore_index=True)
    df = df.dropna(how="all")
    return df, logs
